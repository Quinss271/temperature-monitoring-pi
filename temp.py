#!/usr/bin/python

import python_weather
import asyncio
import os
import Adafruit_DHT
import time
from datetime import datetime
from datetime import date
from openpyxl import Workbook

wb = Workbook()
sensor = Adafruit_DHT.DHT11

ws = wb.active
pin = 23

# Try to grab a sensor reading.  Use the read_retry method which will retry up
# to 15 times to get a sensor reading (waiting 2 seconds between each retry).
humidity, temperature = Adafruit_DHT.read_retry(sensor, pin)

ws['A1'] = "Day"
ws['B1'] = "Month"
ws['C1'] = "Year"
ws['D1'] = "Time"
ws['E1'] = "Inside-C"
ws['F1'] = "humidity-%"
ws['G1'] = "Outside-C"
ws['H1'] = "Diff"
wb.save("temperature3.xlsx")
i=1
while True:
    if humidity is not None and temperature is not None:
            i+=1
            humidity, temperature = Adafruit_DHT.read_retry(sensor, pin)
            print('Temp={0:0.1f}*C  Humidity={1:0.1f}%'.format(temperature, humidity))
            
            now = datetime.now()
            ws['A'+str(i)] = str(now.strftime("%d"))
            ws['B'+str(i)] = str(now.strftime("%m"))
            ws['C'+str(i)] = str(now.strftime("%Y"))
            ws['D'+str(i)] = str(now.strftime("%H:%M:%S"))
            ws['E'+str(i)] = str(temperature)
            ws['F'+str(i)] = str(humidity)
            
            try:
                async def getweather():
                    async with python_weather.Client(format=python_weather.METRIC) as client:
                        weather = await client.get("Melitopol")
                        ws['G'+str(i)] = str(weather.current.temperature)
                        ws['H'+str(i)] = str(temperature-weather.current.temperature)

                if __name__ == "__main__":
                    if os.name == "nt":
                        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
                    asyncio.run(getweather())          
                wb.save("temperature3.xlsx")
            except:
                ws['G'+str(i)] = str("Error")
                ws['H'+str(i)] = str("Error")
                wb.save("temperature3.xlsx")
                print("Error of API")
            time.sleep(1800)

    else:
        print('Failed to get reading. Try again!')
# if __name__ == "__main__":
#   # see https://stackoverflow.com/questions/45600579/asyncio-event-loop-is-closed-when-getting-loop
#   # for more details
#   if os.name == "nt":
#     asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

#   asyncio.run(getweather())
